import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.worksheet.header_footer import HeaderFooter


def create_excel_file(data_liste, file_name):
    # Création d'un nouveau fichier Excel
    wb = Workbook()
    ws = wb.active
    # Configuration du pied de page avec numéros de page
    ws.header_footer = HeaderFooter()
    ws.header_footer.center_footer = "Page &P sur &N"  # Numéro de la page actuelle et total
    ws.title = "Commande"

    for data in data_liste:

        built_tab(data, ws)

    # Définir la largeur des colonnes
    ws.column_dimensions['A'].width = 15  # Largeur pour "N° commande" et "Producteur"
    ws.column_dimensions['B'].width = 40  # Largeur pour "Nom du client" et "Produit"
    ws.column_dimensions['C'].width = 3  # Largeur pour "Quantité"
    ws.column_dimensions['D'].width = 8  # Largeur pour "Prix"

    # Sauvegarde du fichier Excel
    wb.save(file_name)
    print(f"Fichier Excel {file_name} a ete créé avec succès !")


def built_tab(data, ws):
    # Définition des bordures
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Remplissage gris clair
    gris_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Bordure en pointillé gras
    thick_dotted_border = Border(top=Side(style='mediumDashDot'))

    # Ligne 1 : En-têtes pour N° commande et Nom du client
    ws.append(["N° commande", "Nom du client"])

    # Appliquer le remplissage gris clair à la ligne des en-têtes
    for cell in ws[ws.max_row]:
        cell.fill = gris_fill

    # Ligne 2 : Valeurs pour le numéro de commande et le nom du client
    ws.append([data["command"]["numero"], data["command"]["client"]])

    # Appliquer le remplissage gris clair à la ligne des données de commande et client
    for cell in ws[ws.max_row]:
        cell.fill = gris_fill

    # Ligne 3 : En-têtes des colonnes des produits (Producteur, Produit, Quantité, Prix)
    ws.append(["Producteur", "Produit", "Qt", "Prix"])

    # Lignes suivantes : Données des produits
    for produit in data["command"]["produits"]:
        ws.append([
            produit["producteur"],
            produit["produit"],
            produit["quantite"],
            produit["prix"]
        ])

    # Calcul du total
    total = sum([float(p["prix"]) for p in data["command"]["produits"]])

    # Insertion de la ligne Total avec le calcul à la dernière colonne
    ws.append(["Total", "", "", total])

    # Récupérer la cellule contenant "Total" (ligne précédente, colonne 1)
    total_label_cell = ws.cell(row=ws.max_row, column=1)

    # Appliquer la mise en forme gras à la cellule contenant "Total"
    total_label_cell.font = Font(bold=True)

    # Récupérer la cellule contenant le total (ligne précédente, colonne 4)
    total_value_cell = ws.cell(row=ws.max_row, column=4)

    # Appliquer la mise en forme gras à la cellule contenant le total
    total_value_cell.font = Font(bold=True)

    # Appliquer des bordures à chaque cellule du tableau
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    # Ajouter deux lignes vides avec une bordure en pointillé gras entre elles
    ws.append([""])  # Première ligne vide
    ws.append([""])  # Deuxième ligne vide


def create_producer_price_table(df_achat, file_name):
    df_achat['prix'] = pd.to_numeric(df_achat['prix'], errors='coerce')
    total_par_producteur = df_achat.groupby('producteur')['prix'].sum().reset_index()
    liste_totaux = total_par_producteur.to_dict(orient='records')
    # Créer un nouveau fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Producteurs et Prix"

    # Définir les en-têtes des colonnes
    ws.append(["Producteur", "Prix"])

    # Définir des bordures fines pour les cellules
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Ajouter les données dans le tableau
    for item in liste_totaux:
        ws.append([item['producteur'], item['prix']])

    # Appliquer des bordures à chaque cellule
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 30  # Largeur pour "Producteur"
    ws.column_dimensions['B'].width = 15  # Largeur pour "Prix"

    # Sauvegarder le fichier Excel
    wb.save(file_name)
    print(f"Tableau enregistré dans le fichier {file_name} avec succès !")

